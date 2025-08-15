import os
import tempfile
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from models import AcquisitionType

def generate_excel_report(acquisitions):
    """Generate a professional Excel report with acquisition data"""
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Styles
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1e4a6b', end_color='1e4a6b', fill_type='solid')
    title_font = Font(name='Calibri', size=16, bold=True, color='1e4a6b')
    normal_font = Font(name='Calibri', size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: Summary
    summary_sheet = wb.create_sheet("Resumo Executivo")
    
    # Title
    summary_sheet['A1'] = 'SENAI Morvan Figueiredo - Relatório de Aquisições'
    summary_sheet['A1'].font = title_font
    summary_sheet.merge_cells('A1:F1')
    
    summary_sheet['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    summary_sheet['A2'].font = normal_font
    summary_sheet.merge_cells('A2:F2')
    
    # Statistics
    total_acquisitions = len(acquisitions)
    servicos_count = len([a for a in acquisitions if a.type == AcquisitionType.SERVICO])
    insumos_count = len([a for a in acquisitions if a.type == AcquisitionType.INSUMO])
    
    total_value = sum([a.final_value or 0 for a in acquisitions])
    servicos_value = sum([a.final_value or 0 for a in acquisitions if a.type == AcquisitionType.SERVICO])
    insumos_value = sum([a.final_value or 0 for a in acquisitions if a.type == AcquisitionType.INSUMO])
    
    # Summary table
    summary_data = [
        ['Indicador', 'Quantidade', 'Valor (R$)', '', '', ''],
        ['Total de Solicitações', total_acquisitions, total_value, '', '', ''],
        ['Serviços', servicos_count, servicos_value, '', '', ''],
        ['Insumos', insumos_count, insumos_value, '', '', '']
    ]
    
    for row_num, row_data in enumerate(summary_data, 4):
        for col_num, value in enumerate(row_data[:3], 1):
            cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
            cell.font = header_font if row_num == 4 else normal_font
            cell.fill = header_fill if row_num == 4 else PatternFill()
            cell.border = border
            cell.alignment = center_alignment
            
            if col_num == 3 and row_num > 4 and isinstance(value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
    
    # Auto-width columns
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Detailed Data
    details_sheet = wb.create_sheet("Dados Detalhados")
    
    # Create DataFrame
    data = []
    for acquisition in acquisitions:
        data.append({
            'ID': acquisition.id,
            'Título': acquisition.title,
            'Tipo': acquisition.type_display,
            'Categoria': acquisition.category.name if acquisition.category else '',
            'Status': acquisition.status_display,
            'Solicitante': acquisition.requester.full_name,
            'Centro de Custo': acquisition.cost_center.name if acquisition.cost_center else '',
            'Valor Estimado': acquisition.estimated_value or 0,
            'Valor Final': acquisition.final_value or 0,
            'Data Solicitação': acquisition.created_at.strftime('%d/%m/%Y'),
            'Data Aprovação': acquisition.approved_at.strftime('%d/%m/%Y') if acquisition.approved_at else '',
            'Data Conclusão': acquisition.completed_at.strftime('%d/%m/%Y') if acquisition.completed_at else '',
            'Justificativa': acquisition.justification[:100] + ('...' if len(acquisition.justification) > 100 else ''),
            'Fonte da Verba': acquisition.budget_source.value if acquisition.budget_source else '',
            'Método de Pagamento': acquisition.payment_method.value if acquisition.payment_method else ''
        })
    
    df = pd.DataFrame(data)
    
    # Add DataFrame to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        details_sheet.append(r)
    
    # Format header row
    for cell in details_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Format data rows
    for row in details_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Format currency columns
    for row in details_sheet.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value != 0:
                cell.number_format = 'R$ #,##0.00'
    
    # Auto-width columns
    for column in details_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        details_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3: Charts and Analysis
    charts_sheet = wb.create_sheet("Gráficos e Análises")
    
    # Title
    charts_sheet['A1'] = 'Análise Visual dos Dados'
    charts_sheet['A1'].font = title_font
    charts_sheet.merge_cells('A1:H1')
    
    # Type distribution data
    charts_sheet['A3'] = 'Distribuição por Tipo'
    charts_sheet['A3'].font = Font(name='Calibri', size=12, bold=True)
    
    type_data = [
        ['Tipo', 'Quantidade'],
        ['Serviços', servicos_count],
        ['Insumos', insumos_count]
    ]
    
    for row_num, row_data in enumerate(type_data, 4):
        for col_num, value in enumerate(row_data, 1):
            cell = charts_sheet.cell(row=row_num, column=col_num, value=value)
            cell.font = header_font if row_num == 4 else normal_font
            cell.fill = header_fill if row_num == 4 else PatternFill()
            cell.border = border
            cell.alignment = center_alignment
    
    # Pie chart for type distribution
    pie_chart = PieChart()
    labels = Reference(charts_sheet, min_col=1, min_row=5, max_row=6)
    data = Reference(charts_sheet, min_col=2, min_row=4, max_row=6)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.title = "Distribuição: Serviços vs Insumos"
    charts_sheet.add_chart(pie_chart, "D4")
    
    # Status distribution data
    charts_sheet['A10'] = 'Distribuição por Status'
    charts_sheet['A10'].font = Font(name='Calibri', size=12, bold=True)
    
    status_counts = {}
    for acquisition in acquisitions:
        status = acquisition.status_display
        status_counts[status] = status_counts.get(status, 0) + 1
    
    status_data = [['Status', 'Quantidade']]
    for status, count in status_counts.items():
        status_data.append([status, count])
    
    for row_num, row_data in enumerate(status_data, 11):
        for col_num, value in enumerate(row_data, 1):
            cell = charts_sheet.cell(row=row_num, column=col_num, value=value)
            cell.font = header_font if row_num == 11 else normal_font
            cell.fill = header_fill if row_num == 11 else PatternFill()
            cell.border = border
            cell.alignment = center_alignment
    
    # Bar chart for status distribution
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Distribuição por Status"
    bar_chart.y_axis.title = 'Quantidade'
    bar_chart.x_axis.title = 'Status'
    
    data = Reference(charts_sheet, min_col=2, min_row=11, max_row=11+len(status_counts))
    cats = Reference(charts_sheet, min_col=1, min_row=12, max_row=11+len(status_counts))
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    charts_sheet.add_chart(bar_chart, "D12")
    
    # Save file
    wb.save(temp_file.name)
    
    return temp_file.name
